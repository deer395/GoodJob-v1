"""Local CSV/XLSX job-import flow for Phase 2."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import csv
import json
from pathlib import Path
import re
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import RedirectResponse
from fastapi.templating import Jinja2Templates

FIELDS = ("company", "title", "city", "application_url", "salary_range", "deadline", "department", "source", "note")
REQUIRED_FIELDS = ("company", "title", "city")
HEADER_ALIASES = {
    "company": ("公司", "企业", "企业名称", "公司名称", "单位"),
    "title": ("岗位", "职位", "岗位名称", "职位名称", "招聘岗位"),
    "city": ("城市", "工作地点", "工作城市", "地点", "base"),
    "application_url": ("链接", "投递链接", "投递地址", "网址", "官网", "投递链接地址", "简历投递链接", "公告链接"),
    "salary_range": ("薪资", "薪资范围", "薪酬", "待遇"),
    "deadline": ("截止日期", "ddl", "投递截止", "截止时间", "deadline"),
    "department": ("部门", "业务线", "事业部", "事业群"),
    "source": ("来源", "渠道", "数据来源"),
    "note": ("备注", "说明", "其他", "其他信息", "补充"),
}
OVERWRITE_FIELDS = ("company", "title", "city", "application_url", "salary_range", "deadline", "source")
MAX_IMPORT_ROWS = 1000


def _normalized_header(value: str) -> str:
    return "".join((value or "").strip().casefold().split())


def default_mapping(headers: list[str]) -> dict[str, str]:
    normalized_headers = {header: _normalized_header(header) for header in headers}
    mapping = {header: "" for header in headers}
    # One import field must have one source column. This avoids a broad alias
    # such as “链接” silently overwriting a preferred “简历投递链接” later.
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            header = next((item for item in headers if not mapping[item] and normalized_headers[item] == alias), None)
            if header is not None:
                mapping[header] = field
                break
    return mapping


def parse_deadline(value: str, default_year: str) -> tuple[str, str | None]:
    value = (value or "").strip()
    if not value:
        return "", None
    full_match = re.search(r"(\d{4})[./\-年]\s*(\d{1,2})[./\-月]\s*(\d{1,2})", value)
    if full_match:
        try:
            return date(*(int(part) for part in full_match.groups())).isoformat(), None
        except ValueError:
            return "", None
    partial_match = re.fullmatch(r"\s*(\d{1,2})\s*[./月\-]\s*(\d{1,2})\s*(?:日)?\s*", value)
    if partial_match:
        if not default_year:
            return "", "缺少年份"
        try:
            return date(int(default_year), *(int(part) for part in partial_match.groups())).isoformat(), None
        except ValueError:
            return "", None
    # “招满为止”等文本代表无固定 DDL，不应阻断整批导入。
    return "", None


def _read_csv(content: bytes) -> tuple[list[str], list[list[str]], str]:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            rows = list(csv.reader(content.decode(encoding).splitlines()))
            headers, data_rows = _table_rows(rows)
            return headers, data_rows, encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别 CSV 编码，请另存为 UTF-8、GBK 或 GB2312 后重试")


def _header_index(rows: list[list[Any]]) -> int:
    """Find the first real header row instead of assuming the worksheet starts with one."""
    aliases = {_normalized_header(alias) for names in HEADER_ALIASES.values() for alias in names}
    candidates = []
    for index, row in enumerate(rows[:20]):
        values = [str(value or "").strip() for value in row]
        nonempty = sum(bool(value) for value in values)
        matches = sum(_normalized_header(value) in aliases for value in values if value)
        if matches:
            candidates.append((matches, nonempty, -index, index))
    return max(candidates)[-1] if candidates else 0


def _table_rows(rows: list[list[Any]]) -> tuple[list[str], list[list[str]]]:
    if not rows:
        raise ValueError("文件为空")
    header_index = _header_index(rows)
    headers = [str(value).strip() for value in rows[header_index]]
    if not any(headers):
        raise ValueError("文件缺少表头")
    return headers, [["" if value is None else str(value).strip() for value in row[:len(headers)]] for row in rows[header_index + 1:] if any(str(value or "").strip() for value in row)]


def _read_xlsx(content: bytes) -> tuple[list[str], list[list[str]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - package is declared dependency
        raise ValueError("当前环境缺少 openpyxl，请安装依赖后重试") from error
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        candidates = []
        for sheet in workbook.worksheets:
            # Some real-world workbooks have a stale <dimension ref="A1">.
            # Resetting it makes openpyxl stream all rows from sheetData.
            sheet.reset_dimensions()
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            try:
                headers, data_rows = _table_rows(rows)
            except ValueError:
                continue
            # A real data sheet has more records and populated columns than an
            # update log or a cover sheet. This keeps the user's default path
            # simple while still retaining a deterministic selection rule.
            populated_columns = sum(bool(header) for header in headers)
            candidates.append((len(data_rows), populated_columns, sheet.title, headers, data_rows))
        if not candidates:
            raise ValueError("工作簿中没有可识别的带表头工作表")
        _, _, sheet_name, headers, data_rows = max(candidates, key=lambda item: (item[0], item[1]))
        return headers, data_rows, f"xlsx · 工作表：{sheet_name}"
    finally:
        workbook.close()


def _mapped_rows(raw_rows: list[list[str]], headers: list[str], mapping: dict[str, str], default_year: str, store: Any) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for index, raw in enumerate(raw_rows, start=1):
        source_row = dict(zip(headers, raw, strict=False))
        data = {field: "" for field in FIELDS}
        for header, field in mapping.items():
            if field in data:
                data[field] = source_row.get(header, "").strip()
        data["deadline"], date_error = parse_deadline(data["deadline"], default_year)
        key = store.normalized(data["company"], data["title"], data["city"])
        database_duplicate = store.find_duplicate(*key) if all(key) else None
        file_duplicate = key in seen if all(key) else False
        if all(key):
            seen.add(key)
        differences = {}
        if database_duplicate:
            for field in OVERWRITE_FIELDS:
                new_value = data[field]
                old_value = database_duplicate[field] or ""
                if new_value and new_value != old_value:
                    differences[field] = {"old": old_value, "new": new_value}
        missing_fields = [field for field in REQUIRED_FIELDS if not data[field]]
        validation_error = f"缺少{'、'.join(missing_fields)}" if missing_fields else None
        default_action = "skip" if validation_error or file_duplicate or database_duplicate else "create"
        result.append({"row_number": index, "data": data, "date_error": date_error, "validation_error": validation_error, "database_duplicate": dict(database_duplicate) if database_duplicate else None, "file_duplicate": file_duplicate, "differences": differences, "default_action": default_action})
    return result


def create_import_router(templates: Jinja2Templates) -> APIRouter:
    router = APIRouter()

    def render(request: Request, step: str = "upload", **context: Any):
        return templates.TemplateResponse(request, "import.html", {"step": step, "fields": FIELDS, **context})

    @router.get("/import")
    def import_page(request: Request):
        return render(request)

    @router.post("/import/upload")
    async def upload_import(request: Request, file: UploadFile = File(...)):
        filename = Path(file.filename or "").name
        suffix = Path(filename).suffix.lower()
        if suffix == ".xls":
            return render(request, error="请另存为 .xlsx 或 .csv 后重新上传")
        if suffix not in {".csv", ".xlsx"}:
            return render(request, error="仅支持 .csv 与 .xlsx 文件")
        content = await file.read()
        if len(content) > 5 * 1024 * 1024:
            return render(request, error="文件不能超过 5MB")
        try:
            headers, rows, encoding = _read_csv(content) if suffix == ".csv" else _read_xlsx(content)
        except ValueError as error:
            return render(request, error=str(error))
        if len(headers) > 30:
            return render(request, error="文件最多支持 30 列")
        if len(rows) > MAX_IMPORT_ROWS:
            return render(request, error=f"文件最多支持 {MAX_IMPORT_ROWS} 行")
        token = uuid4().hex
        request.app.state.import_cache[token] = {"filename": filename, "headers": headers, "rows": rows, "encoding": encoding}
        return render(request, "mapping", token=token, filename=filename, headers=headers, mapping=default_mapping(headers), sample_rows=rows[:5], default_year=str(date.today().year + 1), encoding=encoding)

    @router.post("/import/preview")
    def preview_import(request: Request, token: str = Form(...), default_year: str = Form(""), mapping_json: str = Form("{}")):
        cached = request.app.state.import_cache.get(token)
        if not cached:
            return render(request, error="导入会话已过期，请重新上传文件")
        try:
            mapping = json.loads(mapping_json)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="列映射格式无效")
        mapping = {header: field for header, field in mapping.items() if header in cached["headers"] and field in FIELDS}
        rows = _mapped_rows(cached["rows"], cached["headers"], mapping, default_year.strip(), request.app.state.store)
        return render(request, "preview", token=token, filename=cached["filename"], mapping_json=json.dumps(mapping, ensure_ascii=False), rows=rows[:10], total_rows=len(rows), invalid_count=sum(bool(row["validation_error"]) for row in rows), default_year=default_year.strip())

    @router.post("/import/execute")
    def execute_import(request: Request, token: str = Form(...), default_year: str = Form(""), mapping_json: str = Form("{}"), actions_json: str = Form("{}")):
        cached = request.app.state.import_cache.get(token)
        if not cached:
            return render(request, error="导入会话已过期，请重新上传文件")
        try:
            mapping, actions = json.loads(mapping_json), json.loads(actions_json)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="导入数据格式无效")
        mapping = {header: field for header, field in mapping.items() if header in cached["headers"] and field in FIELDS}
        rows = _mapped_rows(cached["rows"], cached["headers"], mapping, default_year.strip(), request.app.state.store)
        store = request.app.state.store
        counts = {"created": 0, "skipped": 0, "updated": 0}
        try:
            with store.connection() as conn:
                batch_id = store.create_import_batch(conn, cached["filename"], len(rows), mapping, default_year.strip())
                for row in rows:
                    action = actions.get(str(row["row_number"]), row["default_action"])
                    if action not in {"create", "skip", "update"}:
                        raise ValueError(f"第 {row['row_number']} 行的操作无效")
                    if action == "skip":
                        counts["skipped"] += 1
                        continue
                    if row["date_error"]:
                        raise ValueError(f"第 {row['row_number']} 行：{row['date_error']}")
                    if not all(row["data"][field] for field in REQUIRED_FIELDS):
                        raise ValueError(f"第 {row['row_number']} 行缺少公司、岗位或城市，无法导入")
                    if action == "update":
                        if not row["database_duplicate"]:
                            raise ValueError(f"第 {row['row_number']} 行没有可覆盖的库内记录")
                        store.update_from_import(conn, row["database_duplicate"]["id"], row["data"], batch_id)
                        counts["updated"] += 1
                    else:
                        store.create_from_import(conn, row["data"], batch_id, duplicate_confirmed=bool(row["database_duplicate"] or row["file_duplicate"]))
                        counts["created"] += 1
                store.finish_import_batch(conn, batch_id, counts)
                store.recompute_matches_in_connection(conn)
        except ValueError as error:
            return render(request, "preview", error=str(error), token=token, filename=cached["filename"], mapping_json=json.dumps(mapping, ensure_ascii=False), rows=rows[:10], total_rows=len(rows), invalid_count=sum(bool(row["validation_error"]) for row in rows), default_year=default_year.strip())
        request.app.state.import_cache.pop(token, None)
        return RedirectResponse(url=f"/jobs?imported={counts['created']}&updated={counts['updated']}&skipped={counts['skipped']}", status_code=303)

    @router.get("/import/history")
    def import_history(request: Request):
        return render(request, "history", batches=request.app.state.store.list_import_batches())

    @router.get("/import/history/{batch_id}")
    def import_history_detail(request: Request, batch_id: int):
        batch = request.app.state.store.get_import_batch(batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="导入批次不存在")
        return render(request, "history_detail", batch=batch)

    return router

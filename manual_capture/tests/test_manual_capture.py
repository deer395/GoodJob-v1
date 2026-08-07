from datetime import date, timedelta
from io import BytesIO
import json

import pytest

from fastapi.testclient import TestClient

from manual_capture.app import create_app
from manual_capture.import_routes import default_mapping


def client(tmp_path):
    return TestClient(create_app(tmp_path / "jobs.db"))


def payload(**overrides):
    data = {"company": "字节跳动", "title": "产品经理", "city": "北京", "source": "其他"}
    data.update(overrides)
    return data


def test_saves_job_and_persists_to_sqlite(tmp_path):
    app = create_app(tmp_path / "jobs.db")
    response = TestClient(app).post("/jobs", data=payload(), follow_redirects=True)
    assert response.status_code == 200
    assert "已收录：" in response.text
    assert app.state.store.list()[0]["status"] == "待评估"


def test_required_fields_are_validated(tmp_path):
    response = client(tmp_path).post("/jobs", data=payload(company="", city=""))
    assert "请填写公司名" in response.text
    assert "请填写工作地点" in response.text


def test_duplicate_is_blocked_and_can_be_updated(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload())
    response = test_client.post("/jobs", data=payload(note="重复提交"))
    assert "该岗位已收录" in response.text
    assert "更新原记录" in response.text
    assert len(test_client.app.state.store.list()) == 1


def test_near_and_expired_deadlines_are_shown(tmp_path):
    test_client = client(tmp_path)
    near = (date.today() + timedelta(days=2)).isoformat()
    past = (date.today() - timedelta(days=1)).isoformat()
    assert "临近截止" in test_client.post("/jobs", data=payload(deadline=near), follow_redirects=True).text
    assert "已逾期" in test_client.post("/jobs", data=payload(company="腾讯", deadline=past), follow_redirects=True).text


def test_pool_search_filters_and_sorts_by_deadline(tmp_path):
    test_client = client(tmp_path)
    near = (date.today() + timedelta(days=1)).isoformat()
    later = (date.today() + timedelta(days=10)).isoformat()
    past = (date.today() - timedelta(days=1)).isoformat()
    test_client.post("/jobs", data=payload(company="腾讯", title="后端开发", deadline=later))
    test_client.post("/jobs", data=payload(company="字节跳动", title="产品经理", deadline=near))
    test_client.post("/jobs", data=payload(company="阿里巴巴", title="算法工程师", deadline=past))

    pool = test_client.get("/jobs")
    assert pool.text.index("字节跳动") < pool.text.index("腾讯")
    assert "阿里巴巴" in test_client.get("/jobs?state=expired").text
    assert "腾讯" not in test_client.get("/jobs?state=near").text
    assert "产品经理" in test_client.get("/jobs?q=字节").text
    assert "腾讯" not in test_client.get("/jobs?q=字节").text


def test_note_limit_and_delete_from_edit_page(tmp_path):
    test_client = client(tmp_path)
    too_long = test_client.post("/jobs", data=payload(note="x" * 501))
    assert "备注最多 500 字" in too_long.text

    test_client.post("/jobs", data=payload())
    job_id = test_client.app.state.store.list()[0]["id"]
    deleted = test_client.post(f"/jobs/{job_id}/delete", follow_redirects=True)
    assert deleted.status_code == 200
    assert test_client.app.state.store.list() == []


def test_pool_shows_external_link_only_when_a_job_has_one(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(application_url="https://jobs.example.com/byte"))
    pool = test_client.get("/jobs")
    assert 'href="https://jobs.example.com/byte"' in pool.text
    assert 'target="_blank"' in pool.text


def test_profile_recomputes_explainable_match_and_persists(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(title="后端开发 2027届 硕士", city="北京", department="云产品", note="Python"))
    test_client.post("/profile", data={"graduation_year": "2027", "degree": "硕士", "target_cities": "北京,上海", "target_directions": "后端开发", "skills": "Python"})
    job = test_client.app.state.store.list()[0]
    assert job["match_score"] == 100
    assert "城市匹配" in job["match_reasons"]


def test_graduation_mismatch_is_not_scored_or_high_match(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(title="后端开发 2026届", note="Python"))
    test_client.post("/profile", data={"graduation_year": "2027", "target_cities": "北京", "target_directions": "后端开发", "skills": "Python"})
    job = test_client.app.state.store.list()[0]
    assert job["match_score"] is None
    assert "届别不匹配" in job["match_reasons"]
    assert "字节跳动" not in test_client.get("/jobs?state=high").text


def test_favorite_and_radar_search_include_department(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(department="增长业务"))
    job_id = test_client.app.state.store.list()[0]["id"]
    test_client.post(f"/jobs/{job_id}/favorite")
    assert "字节跳动" in test_client.get("/jobs?state=favorite").text
    assert "字节跳动" in test_client.get("/jobs?q=增长").text


def test_graduation_mismatch_from_job_note_never_receives_a_score(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(company="阿里淘天", city="杭州", note="2027届应届生"))
    test_client.post("/profile", data={"graduation_year": "2026", "target_cities": "杭州"})
    job = test_client.app.state.store.list()[0]
    assert job["match_score"] is None
    assert "届别不匹配" in job["match_reasons"]


def test_direction_atom_match_and_unrelated_direction(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(company="京东", title="产品技术经理", city="北京"))
    test_client.post("/profile", data={"target_directions": "产品经理"})
    assert test_client.app.state.store.list()[0]["match_score"] == 20

    test_client.post("/jobs", data=payload(company="网易", title="前端开发", city="杭州"))
    unmatched = next(job for job in test_client.app.state.store.list() if job["company"] == "网易")
    assert unmatched["match_score"] == 0


def test_char_ngram_threshold_matches_related_words_only(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(company="京东", title="产品技术经理", city="北京"))
    test_client.post("/profile", data={"target_directions": "产品经理"})
    assert test_client.app.state.store.list()[0]["match_score"] == 20


def test_application_workflow_preserves_items_on_withdraw(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload())
    test_client.post("/jobs/1/prepare")
    assert len(test_client.app.state.store.application_items(1)) == 5
    for item in test_client.app.state.store.application_items(1)[:4]:
        test_client.post(f"/checklist/{item['id']}/toggle", data={"app_id": 1})
    test_client.post("/applications/1/save", data={"resume_version":"v1", "next_action":"等待通知", "notes":""})
    assert test_client.app.state.store.confirm_application(1)
    test_client.post("/applications/1/withdraw", data={"confirmed":"true"})
    application = test_client.app.state.store.application(1)
    assert application["status"] == "待投递" and application["applied_at"] is None and application["next_action"] == "等待通知"


def test_confirm_application_accepts_string_form_value_and_next_action_saves(tmp_path):
    test_client = client(tmp_path)
    deadline = (date.today() + timedelta(days=2)).isoformat()
    test_client.post("/jobs", data=payload(deadline=deadline))
    test_client.post("/jobs/1/prepare")
    assert "2 天后截止" in test_client.get("/applications").text

    response = test_client.post("/applications/1/confirm", data={"confirmed": "true"})
    assert response.status_code == 200
    application = test_client.app.state.store.application(1)
    assert application["status"] == "已投递" and application["applied_at"]

    saved = test_client.post("/applications/1/next-action", data={"next_action": "等待笔试通知"})
    assert saved.status_code == 200
    assert test_client.app.state.store.application(1)["next_action"] == "等待笔试通知"


def test_progress_allows_skip_and_creates_events_and_funnel(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(company="字节跳动"))
    test_client.post("/jobs", data=payload(company="腾讯", title="后端开发"))
    test_client.post("/jobs/1/prepare")
    test_client.post("/jobs/2/prepare")
    assert test_client.app.state.store.confirm_application(1)
    assert test_client.app.state.store.confirm_application(2)

    store = test_client.app.state.store
    store.advance_application(1, "面试", "一面", date.today().isoformat(), "一面通知")
    store.advance_application(1, "Offer", "Offer", date.today().isoformat())
    assert store.application_by_id(1)["status"] == "Offer"
    assert {event["event_type"] for event in store.application_events(1)} == {"已投递", "一面", "Offer"}
    assert store.funnel() == {"已投递": 2, "测评/笔试": 0, "面试": 1, "Offer": 1}
    with pytest.raises(ValueError, match="无效的目标阶段|不能回退"):
        store.advance_application(1, "待投递", "已投递", date.today().isoformat())


def test_progress_same_stage_add_event_and_end_statuses(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload())
    test_client.post("/jobs/1/prepare")
    store = test_client.app.state.store
    assert store.confirm_application(1)
    store.advance_application(1, "测评/笔试", "笔试通知", "2026-08-06")
    store.advance_application(1, "测评/笔试", "其他测评", "2026-08-07")
    store.add_application_event(1, "补充材料", "2026-08-08", "材料已提交")
    assert store.application_by_id(1)["status"] == "测评/笔试"
    assert len(store.application_events(1)) == 4
    store.advance_application(1, "已结束", "主动放弃", "2026-08-09")
    assert store.application_by_id(1)["status"] == "已结束"


def test_historical_applied_event_is_backfilled_only_once(tmp_path):
    db_path = tmp_path / "history.db"
    app = create_app(db_path)
    store = app.state.store
    job_id = store.create({**payload(), "application_url": "", "salary_range": "", "department": "", "description_text": "", "deadline": "", "note": ""})
    stamp = "2026-08-01T10:00:00"
    with store.connection() as conn:
        conn.execute("INSERT INTO applications(job_id,status,applied_at,created_at,updated_at) VALUES(?,?,?,?,?)", (job_id, "已投递", stamp, stamp, stamp))
    restarted = create_app(db_path).state.store
    assert len(restarted.application_events(1)) == 1
    create_app(db_path)
    assert len(restarted.application_events(1)) == 1


def test_progress_page_excludes_pending_and_saves_plan_and_withdraws(tmp_path):
    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(company="待投递公司"))
    test_client.post("/jobs/1/prepare")
    assert "待投递公司" not in test_client.get("/progress").text
    assert test_client.app.state.store.confirm_application(1)
    response = test_client.post("/progress/1/next-action", data={"next_action": "准备笔试", "next_action_due_at": "2026-08-10T09:30"}, follow_redirects=True)
    assert "下一步行动已保存" in response.text
    assert test_client.app.state.store.application_by_id(1)["next_action_due_at"] == "2026-08-10T09:30"
    withdrawn = test_client.post("/applications/1/withdraw", data={"confirmed": "true"}, follow_redirects=True)
    assert "已撤回至待投递" in withdrawn.text
    assert "待投递公司" not in test_client.get("/progress").text
    assert test_client.app.state.store.application_events(1)


def test_progress_calendar_projects_existing_events_deadlines_and_plan_time(tmp_path):
    test_client = client(tmp_path)
    today = date.today()
    test_client.post("/jobs", data=payload(deadline=today.isoformat()))
    test_client.post("/jobs/1/prepare")
    store = test_client.app.state.store
    assert store.confirm_application(1)
    store.advance_application(1, "面试", "一面", today.isoformat())
    store.save_application(1, "", "准备面试", "", f"{today.isoformat()}T09:00")

    calendar_page = test_client.get(f"/progress?view=calendar&year={today.year}&month={today.month}")
    assert calendar_page.status_code == 200
    assert "日历视图" in calendar_page.text
    assert "DDL · 字节跳动 · 产品经理" in calendar_page.text
    assert "一面 · 字节跳动 · 产品经理" in calendar_page.text
    assert "下一步 · 字节跳动 · 产品经理" in calendar_page.text


def test_xlsx_import_maps_dates_tracks_batch_and_preserves_protected_fields(tmp_path):
    from openpyxl import Workbook

    test_client = client(tmp_path)
    test_client.post("/jobs", data=payload(department="用户维护部门", note="不可覆盖"))
    existing_id = test_client.app.state.store.list()[0]["id"]
    test_client.post(f"/jobs/{existing_id}/favorite")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["公司", "岗位", "城市", "截止日期", "来源", "链接"])
    sheet.append(["字节跳动", "产品经理", "北京", "8.15", "官网", "https://jobs.example.com/new"])
    sheet.append(["腾讯", "后端开发", "深圳", "2026-08-16", "公众号", ""])
    content = BytesIO()
    workbook.save(content)
    upload = test_client.post("/import/upload", files={"file": ("jobs.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert upload.status_code == 200 and "确认列映射" in upload.text
    token = next(iter(test_client.app.state.import_cache))
    mapping = {"公司": "company", "岗位": "title", "城市": "city", "截止日期": "deadline", "来源": "source", "链接": "application_url"}
    preview = test_client.post("/import/preview", data={"token": token, "default_year": "2026", "mapping_json": json.dumps(mapping, ensure_ascii=False)})
    assert "库内重复" in preview.text and "2026-08-15" in preview.text
    result = test_client.post("/import/execute", data={"token": token, "default_year": "2026", "mapping_json": json.dumps(mapping, ensure_ascii=False), "actions_json": json.dumps({"1": "update", "2": "create"})}, follow_redirects=True)
    assert "新增 1 条、更新 1 条、跳过 0 条" in result.text
    existing = test_client.app.state.store.get(existing_id)
    assert existing["is_favorite"] == 1 and existing["department"] == "用户维护部门" and existing["note"] == "不可覆盖"
    assert existing["application_url"] == "https://jobs.example.com/new"
    batches = test_client.app.state.store.list_import_batches()
    assert len(batches) == 1 and batches[0]["created_count"] == 1 and batches[0]["updated_count"] == 1
    assert all(job["source_import_id"] == batches[0]["id"] for job in test_client.app.state.store.list())


def test_import_marks_missing_year_and_rolls_back_invalid_rows(tmp_path):
    test_client = client(tmp_path)
    source = "公司,岗位,城市,DDL\n网易,算法工程师,杭州,8.15\n"
    test_client.post("/import/upload", files={"file": ("jobs.csv", source.encode(), "text/csv")})
    token = next(iter(test_client.app.state.import_cache))
    mapping = {"公司": "company", "岗位": "title", "城市": "city", "DDL": "deadline"}
    preview = test_client.post("/import/preview", data={"token": token, "default_year": "", "mapping_json": json.dumps(mapping, ensure_ascii=False)})
    assert "缺少年份" in preview.text
    response = test_client.post("/import/execute", data={"token": token, "default_year": "", "mapping_json": json.dumps(mapping, ensure_ascii=False), "actions_json": "{}"})
    assert "第 1 行：缺少年份" in response.text
    assert test_client.app.state.store.list() == [] and test_client.app.state.store.list_import_batches() == []


def test_xlsx_import_selects_data_sheet_and_skips_preface_rows(tmp_path):
    from openpyxl import Workbook

    test_client = client(tmp_path)
    workbook = Workbook()
    cover = workbook.active
    cover.title = "更新说明"
    cover.append(["更新时间"])
    cover.append(["2026-08-06"])
    sheet = workbook.create_sheet("校招汇总表")
    sheet.append(["2027 秋招汇总，持续更新"])
    sheet.append([])
    sheet.append(["更新时间", "公司", "岗位", "城市", "截止日期"])
    sheet.append(["2026-08-06", "腾讯", "产品经理", "深圳", "2026/08/20"])
    sheet.append(["2026-08-06", "网易", "算法工程师", "杭州", "8.21"])
    content = BytesIO()
    workbook.save(content)
    response = test_client.post("/import/upload", files={"file": ("realistic.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert "工作表：校招汇总表" in response.text
    token = next(iter(test_client.app.state.import_cache))
    cached = test_client.app.state.import_cache[token]
    assert cached["headers"] == ["更新时间", "公司", "岗位", "城市", "截止日期"]
    assert len(cached["rows"]) == 2


def test_xlsx_import_resets_stale_dimension_and_accepts_large_local_batch(tmp_path):
    from openpyxl import Workbook

    test_client = client(tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["公司", "招聘岗位", "工作地点", "截止时间", "简历投递链接"])
    for number in range(659):
        sheet.append([f"公司{number}", "产品经理", "上海", "招满为止" if number % 2 else "2026/08/20", "https://jobs.example.com/apply"])
    content = BytesIO()
    workbook.save(content)
    response = test_client.post("/import/upload", files={"file": ("large.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert response.status_code == 200 and "确认列映射" in response.text
    token = next(iter(test_client.app.state.import_cache))
    cached = test_client.app.state.import_cache[token]
    assert len(cached["rows"]) == 659
    assert cached["headers"] == ["公司", "招聘岗位", "工作地点", "截止时间", "简历投递链接"]
    assert "application_url" in default_mapping(cached["headers"]).values()


def test_import_skips_invalid_rows_outside_the_first_ten_previewed_rows(tmp_path):
    test_client = client(tmp_path)
    rows = ["公司,岗位,城市"] + [f"公司{index},产品经理,上海" for index in range(11)] + ["缺岗位公司,,北京"]
    test_client.post("/import/upload", files={"file": ("invalid-late.csv", "\n".join(rows).encode(), "text/csv")})
    token = next(iter(test_client.app.state.import_cache))
    mapping = {"公司": "company", "岗位": "title", "城市": "city"}
    preview = test_client.post("/import/preview", data={"token": token, "default_year": "2027", "mapping_json": json.dumps(mapping, ensure_ascii=False)})
    assert "其中 1 行缺少公司、岗位或城市，已默认跳过" in preview.text
    completed = test_client.post("/import/execute", data={"token": token, "default_year": "2027", "mapping_json": json.dumps(mapping, ensure_ascii=False), "actions_json": "{}"}, follow_redirects=True)
    assert "新增 11 条、更新 0 条、跳过 1 条" in completed.text


def test_calendar_export_includes_only_eligible_entries_and_uses_crlf(tmp_path):
    test_client = client(tmp_path)
    future = (date.today() + timedelta(days=3)).isoformat()
    test_client.post("/jobs", data=payload(company="待投递", deadline=future))
    test_client.post("/jobs", data=payload(company="已投递", title="后端", deadline=future))
    test_client.post("/jobs", data=payload(company="面试公司", title="算法", deadline=future))
    store = test_client.app.state.store
    test_client.post("/jobs/2/prepare")
    assert store.confirm_application(1)
    test_client.post("/jobs/3/prepare")
    assert store.confirm_application(2)
    store.advance_application(2, "面试", "一面", future, scheduled_at=f"{future}T14:00")
    store.add_application_event(2, "二面", future, scheduled_at="")
    store.save_application(2, "", "准备面试", "", f"{future}T09:00")
    exported = test_client.get("/calendar/export?scope=future")
    body = exported.content.decode()
    assert exported.headers["content-type"].startswith("text/calendar")
    assert "DDL：待投递 产品经理" in body and "DDL：已投递 后端" not in body
    assert "一面：面试公司 算法" in body and "二面：面试公司 算法" not in body
    assert "下一步：面试公司 算法" in body and "UID:campusai-event-" in body
    assert "\r\n" in body and "\n" not in body.replace("\r\n", "")


def test_calendar_export_excludes_notifications_applied_events_and_ended_applications(tmp_path):
    test_client = client(tmp_path)
    future = (date.today() + timedelta(days=4)).isoformat()
    test_client.post("/jobs", data=payload(company="通知事件", deadline=future))
    test_client.post("/jobs", data=payload(company="已结束事件", title="测试", deadline=future))
    store = test_client.app.state.store
    test_client.post("/jobs/1/prepare")
    assert store.confirm_application(1)
    store.add_application_event(1, "笔试通知", future, scheduled_at=f"{future}T10:00")
    store.add_application_event(1, "已投递", future, scheduled_at=f"{future}T11:00")
    test_client.post("/jobs/2/prepare")
    assert store.confirm_application(2)
    store.advance_application(2, "面试", "一面", future, scheduled_at=f"{future}T14:00")
    store.save_application(2, "", "已结束后不应导出", "", f"{future}T09:00")
    store.advance_application(2, "已结束", "主动放弃", future)

    body = test_client.get("/calendar/export?scope=future").content.decode()
    assert "笔试通知：通知事件 产品经理" not in body
    assert "已投递：通知事件 产品经理" not in body
    assert "一面：已结束事件 测试" not in body
    assert "下一步：已结束事件 测试" not in body
